#!/usr/bin/env python3
"""
Professional Sprint Planning Excel Generator
============================================

This script generates a comprehensive, color-coded Excel workbook
for sprint planning with multiple data views.

Requirements:
    pip install openpyxl

Usage:
    python sprint_planner_generator.py

Author: Sahil Betal
Date: March 2026
Project: UIDAI Data Hackathon - Aadhaar Trends
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
import os


class SprintPlannerGenerator:
    """Generate professional sprint planning Excel workbooks"""
    
    def __init__(self):
        """Initialize color palette and styles"""
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        
        # Professional color palette - Strong colors for better visibility
        self.COLORS = {
            'P0_dark': 'C00000',      # Dark red for P0
            'P0_light': 'C00000',     # Strong red (same as dark)
            'P1_dark': 'E67E22',      # Dark orange for P1
            'P1_light': 'E67E22',     # Strong orange (same as dark)
            'P2_dark': 'F39C12',      # Dark gold for P2
            'P2_light': 'F39C12',     # Strong gold (same as dark)
            'P3_dark': '27AE60',      # Dark green for P3
            'P3_light': '27AE60',     # Strong green (same as dark)
            'header': '1C3A5C',       # Very dark blue header
            'subheader': '2E5A8C',    # Dark blue subheader
            'text': 'FFFFFF',
            'light_gray': '34495E',   # Dark gray
            'border': '000000',
            'blocked': 'CC0000',      # Bright red for blocked
            'done': '16A34A',         # Strong green for done
            'in_progress': '0284C7',  # Strong blue for in progress
            'not_started': 'DC2626',  # Bright red for not started
        }
        
        self.border = Border(
            left=Side(style='thin', color=self.COLORS['border']),
            right=Side(style='thin', color=self.COLORS['border']),
            top=Side(style='thin', color=self.COLORS['border']),
            bottom=Side(style='thin', color=self.COLORS['border'])
        )
    
    def get_priority_color(self, priority):
        """Return dark and light colors based on priority"""
        color_map = {
            'P0': (self.COLORS['P0_dark'], self.COLORS['P0_light']),
            'P1': (self.COLORS['P1_dark'], self.COLORS['P1_light']),
            'P2': (self.COLORS['P2_dark'], self.COLORS['P2_light']),
            'P3': (self.COLORS['P3_dark'], self.COLORS['P3_light']),
        }
        return color_map.get(priority, (self.COLORS['P3_dark'], self.COLORS['P3_light']))
    
    def create_header_row(self, ws, row, headers, col_widths=None):
        """Create a formatted header row"""
        header_fill = PatternFill(
            start_color=self.COLORS['header'],
            end_color=self.COLORS['header'],
            fill_type='solid'
        )
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Set column widths if provided
        if col_widths:
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
        
        # Create header cells
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = self.border
        
        ws.row_dimensions[row].height = 25
    
    def create_sprint_master_plan(self):
        """Create Sprint Master Plan sheet with 100 user stories"""
        print("Creating Sheet 1: Sprint Master Plan...")
        
        ws = self.wb.create_sheet("Sprint Master Plan")
        
        # Column widths
        col_widths = [12, 20, 15, 10, 12, 12, 15, 12, 15, 15, 12, 15, 15, 12, 15, 15]
        
        # Headers
        headers = [
            'Story ID', 'Title', 'Epic', 'Priority', 'Story Points',
            'Sprint', 'Owner', 'Risk Level', 'Start Date', 'End Date', 
            'Est. Days', 'Progress %', 'Blockers', 'Status', 'Dependencies', 'Notes'
        ]
        
        self.create_header_row(ws, 1, headers, col_widths)
        
        # Generate 100 user stories
        priorities = ['P0', 'P0', 'P0', 'P1', 'P1', 'P2', 'P3']
        sprints = ['Sprint 1', 'Sprint 2', 'Sprint 3', 'Sprint 4', 'Sprint 5']
        epics = ['Data Setup', 'EDA', 'Data Processing', 'Modeling', 
                 'Insights', 'Visualization', 'Documentation', 'QA']
        owners = ['Sahil Betal', 'Anjali Kumar', 'Maithili Singh', 'Vijay Patel']
        risk_levels = ['Low', 'Medium', 'High', 'Critical']
        statuses = ['Not Started', 'In Progress', 'In Review', 'Done']
        
        base_date = datetime(2026, 3, 18)
        
        for i in range(1, 101):
            priority = random.choice(priorities)
            sprint = random.choice(sprints)
            epic = random.choice(epics)
            owner = random.choice(owners)
            risk_level = random.choice(risk_levels)
            status = random.choice(statuses)
            points = random.choice([3, 5, 8, 13, 21])
            days = random.randint(1, 5)
            progress = random.randint(0, 100) if status != 'Not Started' else 0
            blockers = random.choice(['None', 'Waiting for input', 'Dependencies pending', 'To Be Determined'])
            
            # Generate realistic dependencies
            dependency_choices = [
                'None',
                f'Depends on US-{max(1, i - random.randint(1, 10)):03d}',
                f'Depends on US-{max(1, i - random.randint(1, 15)):03d}, US-{max(1, i - random.randint(1, 20)):03d}',
                'Blocked by API design',
                'Blocked by infrastructure setup',
                'Blocked by data availability',
                'Blocked by requirements finalization',
                f'Depends on {epic} completion',
                'Awaiting stakeholder approval',
                'Blocked by external dependency'
            ]
            dependencies = random.choice(dependency_choices)
            
            start_date = base_date + timedelta(days=random.randint(0, 20))
            end_date = start_date + timedelta(days=days)
            
            data = [
                f'US-{i:03d}',
                f'User Story {i}',
                epic,
                priority,
                points,
                sprint,
                owner,
                risk_level,
                start_date.strftime('%m/%d/%Y'),
                end_date.strftime('%m/%d/%Y'),
                days,
                f'{progress}%',
                blockers,
                status,
                dependencies,
                f'Notes for US-{i:03d}'
            ]
            
            row = i + 1
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Priority coloring
                if col == 4:
                    dark_color, light_color = self.get_priority_color(priority)
                    cell.fill = PatternFill(start_color=dark_color, end_color=dark_color, fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF', size=10)
                
                # Owner coloring
                if col == 7:
                    owner_colors = {
                        'Sahil Betal': '16A34A',
                        'Anjali Kumar': 'D97706',
                        'Maithili Singh': '0284C7',
                        'Vijay Patel': '7C3AED'
                    }
                    if value in owner_colors:
                        cell.fill = PatternFill(start_color=owner_colors[value], end_color=owner_colors[value], fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF', size=10)
                
                # Risk Level coloring
                if col == 8:
                    risk_colors = {
                        'Low': '16A34A',
                        'Medium': 'F59E0B',
                        'High': 'E67E22',
                        'Critical': 'CC0000'
                    }
                    if value in risk_colors:
                        cell.fill = PatternFill(start_color=risk_colors[value], end_color=risk_colors[value], fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF', size=10)
                
                # Progress coloring
                if col == 12:
                    if progress == 100:
                        color = '16A34A'
                    elif progress >= 75:
                        color = '84CC16'
                    elif progress >= 50:
                        color = 'F59E0B'
                    else:
                        color = 'DC2626'
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF', size=10)
                
                # Status coloring
                if col == 14:
                    status_colors = {
                        'Done': '16A34A',
                        'In Review': 'F59E0B',
                        'In Progress': '0284C7',
                        'Not Started': 'DC2626'
                    }
                    if status in status_colors:
                        cell.fill = PatternFill(
                            start_color=status_colors[status],
                            end_color=status_colors[status],
                            fill_type='solid'
                        )
                        cell.font = Font(bold=True, color='FFFFFF', size=10)
                
                # Dependencies coloring
                if col == 15:
                    if 'None' in value:
                        color = '16A34A'
                    elif 'Blocked' in value:
                        color = 'CC0000'
                    elif 'Depends on' in value:
                        color = '0284C7'
                    elif 'Awaiting' in value:
                        color = 'F59E0B'
                    else:
                        color = '7C3AED'
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF', size=9)
            
            ws.row_dimensions[row].height = 20
        
        ws.freeze_panes = 'A2'
        print(f"  ✓ Added 100 user stories to Sprint Master Plan with team assignments")
    
    def create_team_capacity(self):
        """Create Team Capacity sheet"""
        print("Creating Sheet 2: Team Capacity & Allocation...")
        
        ws = self.wb.create_sheet("Team Capacity")
        
        col_widths = [18, 18, 12, 12, 15, 15, 15, 15, 12, 18]
        
        # Title
        title_cell = ws['A1']
        title_cell.value = 'TEAM CAPACITY & RESOURCE ALLOCATION'
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(
            start_color=self.COLORS['header'],
            end_color=self.COLORS['header'],
            fill_type='solid'
        )
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:J1')
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = [
            'Team Member', 'Role', 'Sprint 1', 'Sprint 2', 'Sprint 3',
            'Sprint 4', 'Sprint 5', 'Total Points', 'Utilization %', 'Status'
        ]
        
        self.create_header_row(ws, 2, headers, col_widths)
        
        # Team members with their assignments
        team_members = [
            ('Sahil Betal', 'Tech Lead / Data Scientist', 45, 35, 40, 45, 50),
            ('Anjali Kumar', 'Data Engineer', 40, 47, 30, 35, 60),
            ('Maithili Singh', 'ML Engineer', 30, 25, 44, 40, 35),
            ('Vijay Patel', 'Full Stack Developer', 35, 40, 35, 51, 40),
        ]
        
        row = 3
        for name, role, s1, s2, s3, s4, s5 in team_members:
            total = s1 + s2 + s3 + s4 + s5
            utilization = min(100, (total / 200) * 100)
            status = 'Available' if utilization < 80 else 'Allocated' if utilization < 95 else 'At Capacity'
            
            data = [name, role, s1, s2, s3, s4, s5, total, f'{utilization:.1f}%', status]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10)
                
                # Team member name coloring
                if col == 1:
                    owner_colors = {
                        'Sahil Betal': '16A34A',
                        'Anjali Kumar': 'D97706',
                        'Maithili Singh': '0284C7',
                        'Vijay Patel': '7C3AED'
                    }
                    if value in owner_colors:
                        cell.fill = PatternFill(start_color=owner_colors[value], end_color=owner_colors[value], fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF', size=10)
                
                # Utilization coloring
                if col == 9:
                    if utilization >= 95:
                        color = 'CC0000'
                    elif utilization >= 80:
                        color = 'F59E0B'
                    else:
                        color = '16A34A'
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF', size=10)
                
                # Status coloring
                if col == 10:
                    status_colors = {
                        'Available': '16A34A',
                        'Allocated': 'F59E0B',
                        'At Capacity': 'CC0000'
                    }
                    if value in status_colors:
                        cell.fill = PatternFill(start_color=status_colors[value], end_color=status_colors[value], fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF', size=10)
            
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Totals row
        totals = ['TOTAL', '', 150, 147, 149, 171, 185, 802, '']
        for col, value in enumerate(totals, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.fill = PatternFill(
                start_color='34495E',
                end_color='34495E',
                fill_type='solid'
            )
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        ws.row_dimensions[row].height = 20
        print(f"  ✓ Added team capacity tracking for {len(team_members)} members")
    
    def create_daily_tasks(self):
        """Create Daily Tasks sheet with 120 tasks"""
        print("Creating Sheet 3: Daily Task Breakdown...")
        
        ws = self.wb.create_sheet("Daily Tasks")
        
        col_widths = [12, 12, 20, 18, 15, 12, 15, 15, 12, 20]
        
        # Title
        title_cell = ws['A1']
        title_cell.value = 'DAILY TASK BREAKDOWN - 15 DAY SPRINT'
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(
            start_color=self.COLORS['header'],
            end_color=self.COLORS['header'],
            fill_type='solid'
        )
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:J1')
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = ['Day', 'Date', 'Task', 'Owner', 'Story ID', 'Hours', 'Priority', 'Status', 'Completed %', 'Notes']
        self.create_header_row(ws, 2, headers, col_widths)
        
        # Tasks
        tasks = [
            'Data exploration', 'Data cleaning', 'Feature engineering', 'Model training',
            'Hyperparameter tuning', 'Visualization', 'Dashboard development', 'Documentation'
        ]
        
        owners = ['Sahil Betal', 'Anjali Kumar', 'Maithili Singh', 'Vijay Patel']
        priorities = ['P0', 'P1', 'P2', 'P3']
        
        row = 3
        current_date = datetime(2026, 3, 18)
        
        for day in range(1, 16):
            for task_idx, task in enumerate(tasks):
                owner = owners[task_idx % len(owners)]
                priority = random.choice(priorities)
                status = random.choice(['Not Started', 'In Progress', 'Done'])
                completed = random.randint(0, 100) if status != 'Not Started' else 0
                
                data = [
                    day,
                    current_date.strftime('%m/%d'),
                    task,
                    owner,
                    f'US-{random.randint(1, 100):03d}',
                    random.choice([2, 4, 6, 8]),
                    priority,
                    status,
                    f'{completed}%',
                    f'Task details for Day {day}'
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(size=10)
                    
                    # Owner coloring
                    if col == 4:
                        owner_colors = {
                            'Sahil Betal': '16A34A',
                            'Anjali Kumar': 'D97706',
                            'Maithili Singh': '0284C7',
                            'Vijay Patel': '7C3AED'
                        }
                        if value in owner_colors:
                            cell.fill = PatternFill(start_color=owner_colors[value], end_color=owner_colors[value], fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFF', size=10)
                    
                    # Priority coloring
                    if col == 7:
                        priority_colors = {
                            'P0': 'C00000',
                            'P1': 'E67E22',
                            'P2': 'F39C12',
                            'P3': '27AE60'
                        }
                        if value in priority_colors:
                            cell.fill = PatternFill(start_color=priority_colors[value], end_color=priority_colors[value], fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFF', size=10)
                    
                    # Status coloring
                    if col == 8:
                        status_colors = {
                            'Done': '16A34A',
                            'In Progress': '0284C7',
                            'Not Started': 'DC2626'
                        }
                        if value in status_colors:
                            cell.fill = PatternFill(
                                start_color=status_colors[value],
                                end_color=status_colors[value],
                                fill_type='solid'
                            )
                            cell.font = Font(bold=True, color='FFFFFF', size=10)
                    
                    # Completed % coloring
                    if col == 9:
                        if completed == 100:
                            color = '16A34A'
                        elif completed >= 75:
                            color = '84CC16'
                        elif completed >= 25:
                            color = '0284C7'
                        else:
                            color = 'DC2626'
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF', size=10)
                
                ws.row_dimensions[row].height = 18
                row += 1
            
            current_date += timedelta(days=1)
        
        ws.freeze_panes = 'A3'
        print(f"  ✓ Added 120 daily tasks (15 days × 8 tasks) with team assignments")
    
    def create_burndown_data(self):
        """Create Burndown Data sheet"""
        print("Creating Sheet 4: Sprint Burndown Tracking...")
        
        ws = self.wb.create_sheet("Burndown Data")
        
        col_widths = [15, 18, 18, 18, 18, 18, 15]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        current_row = 1
        sprint_targets = [45, 47, 44, 51, 68]
        
        for sprint_num, target_points in enumerate(sprint_targets, 1):
            sprint_name = f'Sprint {sprint_num}'
            
            # Sprint title
            title_cell = ws.cell(row=current_row, column=1)
            title_cell.value = sprint_name.upper()
            title_cell.font = Font(bold=True, size=12, color='FFFFFF')
            title_cell.fill = PatternFill(
                start_color=self.COLORS['header'],
                end_color=self.COLORS['header'],
                fill_type='solid'
            )
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
            # Headers
            headers = ['Day', 'Planned Points', 'Completed Points', 'Remaining Points', 
                      'Velocity', 'Team Capacity', 'Notes']
            
            header_fill = PatternFill(
                start_color=self.COLORS['subheader'],
                end_color=self.COLORS['subheader'],
                fill_type='solid'
            )
            header_font = Font(bold=True, color='FFFFFF', size=10)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
            
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            
            # Burndown data (3 days per sprint)
            remaining = target_points
            for day in range(1, 4):
                completed = int(target_points * day / 3 * random.uniform(0.8, 1.1))
                remaining = max(0, target_points - completed)
                velocity = completed / day
                
                data = [
                    f'Day {day}',
                    target_points,
                    completed,
                    remaining,
                    f'{velocity:.1f}',
                    '40 hrs',
                    f'Burndown for {sprint_name} Day {day}'
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = value
                    cell.border = self.border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(size=10)
                    
                    # Remaining points coloring
                    if col == 4:
                        if remaining > target_points * 0.3:
                            color = 'CC0000'
                        elif remaining > 0:
                            color = 'F59E0B'
                        else:
                            color = '16A34A'
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF', size=10)
                
                ws.row_dimensions[current_row].height = 18
                current_row += 1
            
            current_row += 1
        
        print(f"  ✓ Added burndown tracking for all 5 sprints")
    
    def create_priority_matrix(self):
        """Create Priority Matrix sheet (90 rows × 20 columns)"""
        print("Creating Sheet 5: Priority Matrix...")
        
        ws = self.wb.create_sheet("Priority Matrix")
        
        # Set column widths
        for i in range(1, 21):
            ws.column_dimensions[get_column_letter(i)].width = 12
        
        # Title
        title_cell = ws['A1']
        title_cell.value = 'PRIORITY × IMPACT MATRIX'
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(
            start_color=self.COLORS['header'],
            end_color=self.COLORS['header'],
            fill_type='solid'
        )
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:T1')
        ws.row_dimensions[1].height = 25
        
        # Matrix headers
        priorities = ['P0', 'P0', 'P0', 'P0', 'P0', 'P1', 'P1', 'P1', 'P1', 'P1',
                     'P2', 'P2', 'P2', 'P2', 'P2', 'P3', 'P3', 'P3', 'P3', 'P3']
        impacts = ['Critical', 'High', 'Medium', 'Low', 'Minimal'] * 4
        
        for col, (priority, impact) in enumerate(zip(priorities, impacts), 1):
            cell = ws.cell(row=2, column=col)
            cell.value = f'{priority}\n{impact}'
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            
            dark_color, light_color = self.get_priority_color(priority)
            cell.fill = PatternFill(start_color=dark_color, end_color=dark_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.border
        
        ws.row_dimensions[2].height = 25
        
        # Generate 90 data rows
        for row in range(3, 93):
            for col in range(1, 21):
                cell = ws.cell(row=row, column=col)
                story_id = f'US-{random.randint(1, 100):03d}'
                points = random.choice([3, 5, 8, 13, 21])
                cell.value = f'{story_id}\n({points}pts)'
                
                dark_color, light_color = self.get_priority_color(priorities[col - 1])
                cell.fill = PatternFill(start_color=dark_color, end_color=dark_color, fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF', size=8)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.border
            
            ws.row_dimensions[row].height = 25
        
        print(f"  ✓ Added priority matrix with 90 rows × 20 columns (1800 cells)")
    
    def create_epic_roadmap(self):
        """Create Epic Roadmap sheet"""
        print("Creating Sheet 6: Epic Roadmap...")
        
        ws = self.wb.create_sheet("Epic Roadmap")
        
        col_widths = [15, 30, 12, 15, 15, 15, 15, 20, 15]
        
        # Title
        title_cell = ws['A1']
        title_cell.value = 'EPIC ROADMAP & TIMELINE'
        title_cell.font = Font(bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill(
            start_color=self.COLORS['header'],
            end_color=self.COLORS['header'],
            fill_type='solid'
        )
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 25
        
        # Headers
        headers = ['Epic ID', 'Epic Name', 'Stories', 'Total Points', 
                  'Sprint 1', 'Sprint 2', 'Sprint 3', 'Sprint 4', 'Sprint 5']
        self.create_header_row(ws, 2, headers, col_widths)
        
        # Epic data
        epics_data = [
            ('EP-001', 'Data Setup', 4, 12, 12, 0, 0, 0, 0),
            ('EP-002', 'EDA', 7, 63, 33, 30, 0, 0, 0),
            ('EP-003', 'Data Processing', 6, 58, 0, 17, 41, 0, 0),
            ('EP-004', 'Modeling', 6, 62, 0, 0, 3, 59, 0),
            ('EP-005', 'Insights', 6, 45, 0, 0, 0, 21, 24),
            ('EP-006', 'Visualization', 2, 29, 0, 0, 0, 0, 29),
            ('EP-007', 'Documentation', 4, 42, 0, 0, 0, 0, 42),
            ('EP-008', 'QA', 3, 26, 0, 0, 0, 0, 26),
        ]
        
        row = 3
        for epic_id, epic_name, stories, total, s1, s2, s3, s4, s5 in epics_data:
            data = [epic_id, epic_name, stories, total, s1, s2, s3, s4, s5]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10)
                
                # Color sprint columns
                if col >= 5 and value > 0:
                    cell.fill = PatternFill(
                        start_color='E67E22',
                        end_color='E67E22',
                        fill_type='solid'
                    )
                    cell.font = Font(bold=True, color='FFFFFF', size=10)
            
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Totals row
        total_row = row
        totals = ['TOTAL', '', sum([d[2] for d in epics_data]), 
                 sum([d[3] for d in epics_data]),
                 sum([d[4] for d in epics_data]), sum([d[5] for d in epics_data]),
                 sum([d[6] for d in epics_data]), sum([d[7] for d in epics_data]),
                 sum([d[8] for d in epics_data])]
        
        for col, value in enumerate(totals, 1):
            cell = ws.cell(row=total_row, column=col)
            cell.value = value
            cell.fill = PatternFill(
                start_color='34495E',
                end_color='34495E',
                fill_type='solid'
            )
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        ws.row_dimensions[total_row].height = 20
        print(f"  ✓ Added epic roadmap with {len(epics_data)} epics")
    
    def generate(self, filename="Professional_Sprint_Planner.xlsx"):
        """Generate the complete workbook"""
        print("\n" + "=" * 80)
        print("🚀 Generating Professional Sprint Planning Excel Workbook...")
        print("=" * 80 + "\n")
        
        # Create all sheets
        self.create_sprint_master_plan()
        self.create_team_capacity()
        self.create_daily_tasks()
        self.create_burndown_data()
        self.create_priority_matrix()
        self.create_epic_roadmap()
        
        # Save workbook
        self.wb.save(filename)
        
        file_size = os.path.getsize(filename) / 1024
        
        print("\n" + "=" * 80)
        print(f"✅ EXCEL WORKBOOK GENERATED SUCCESSFULLY!")
        print("=" * 80)
        print(f"\nFile: {filename}")
        print(f"Size: {file_size:.1f} KB")
        print(f"\nSheets Created:")
        print(f"  1. Sprint Master Plan      (100 rows × 14 columns)")
        print(f"  2. Team Capacity          (5 team members + totals)")
        print(f"  3. Daily Tasks            (120 tasks - 15 days × 8 tasks)")
        print(f"  4. Burndown Data          (5 sprints × 3 days + headers)")
        print(f"  5. Priority Matrix        (90 rows × 20 columns = 1800 cells)")
        print(f"  6. Epic Roadmap           (8 epics + sprint breakdown)")
        print(f"\nColor Scheme:")
        print(f"  🔴 P0 (Red)    - Critical priority")
        print(f"  🟠 P1 (Orange) - High priority")
        print(f"  🟡 P2 (Gold)   - Medium priority")
        print(f"  🟢 P3 (Green)  - Low priority")
        print(f"\nFeatures:")
        print(f"  ✓ Professional formatting")
        print(f"  ✓ Color-coded priorities")
        print(f"  ✓ Borders and alignment")
        print(f"  ✓ Multiple sheets for different views")
        print(f"  ✓ Frozen headers")
        print(f"  ✓ Ready for immediate use")
        print(f"\n✨ Total data cells: 1000+ with professional styling\n")
        
        return filename


if __name__ == "__main__":
    # Generate the workbook
    generator = SprintPlannerGenerator()
    output_file = generator.generate()
    
    print(f"📁 File saved to: {os.path.abspath(output_file)}")
    print("\n✨ Ready to use! Open the file in Excel and start planning.\n")