# ==========================================================
# AUTO-GENERATED SPRINT PLAN FROM PRODUCT BACKLOG
# ==========================================================
# pip install openpyxl pandas

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ==========================
# CONFIGURATION
# ==========================
BACKLOG_FILE = "Product-Backlog.xlsx"
SPRINT_VELOCITY = 20   # Max Story Points per Sprint
SPRINT_LENGTH_DAYS = 10

# ==========================
# LOAD PRODUCT BACKLOG
# ==========================
df = pd.read_excel(BACKLOG_FILE)

# Normalize priority for sorting
priority_map = {"High": 1, "Medium": 2, "Low": 3}
df["PriorityRank"] = df["Priority"].map(priority_map)

# Sort by priority first, then story points (descending impact)
df = df.sort_values(by=["PriorityRank", "Story Points"], ascending=[True, False])

# ==========================
# AUTO SPRINT ALLOCATION
# ==========================
current_sprint = 1
current_points = 0
sprint_assignments = []

for _, row in df.iterrows():
    sp = row["Story Points"]

    if current_points + sp > SPRINT_VELOCITY:
        current_sprint += 1
        current_points = 0

    sprint_assignments.append(current_sprint)
    current_points += sp

df["Sprint No"] = sprint_assignments

# ==========================
# CREATE WORKBOOK
# ==========================
wb = Workbook()

# ==========================
# SPRINT PLAN SHEET
# ==========================
ws_plan = wb.active
ws_plan.title = "Sprint Plan"

headers = ["Sprint No", "User Story ID", "Description",
           "Priority", "Story Points", "Status"]

ws_plan.append(headers)

for col in range(1, len(headers)+1):
    ws_plan.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
    ws_plan.cell(row=1, column=col).fill = PatternFill("solid", fgColor="1F4E78")
    ws_plan.column_dimensions[get_column_letter(col)].width = 22

for _, row in df.iterrows():
    ws_plan.append([
        row["Sprint No"],
        row["User Story ID"],
        row["Description"],
        row["Priority"],
        row["Story Points"],
        row["Status"]
    ])

# ==========================
# SPRINT METRICS
# ==========================
ws_metrics = wb.create_sheet("Sprint Metrics")

metrics_headers = ["Sprint No", "Planned SP", "Completed SP", "Velocity"]
ws_metrics.append(metrics_headers)

for col in range(1, len(metrics_headers)+1):
    ws_metrics.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
    ws_metrics.cell(row=1, column=col).fill = PatternFill("solid", fgColor="244062")
    ws_metrics.column_dimensions[get_column_letter(col)].width = 22

max_sprint = df["Sprint No"].max()

for sprint in range(1, max_sprint+1):
    ws_metrics.append([
        sprint,
        f"=SUMIF('Sprint Plan'!A:A,{sprint},'Sprint Plan'!E:E)",
        f"=SUMIFS('Sprint Plan'!E:E,'Sprint Plan'!A:A,{sprint},'Sprint Plan'!F:F,\"Done\")",
        f"=C{sprint+1}"
    ])

# Velocity Chart
chart = BarChart()
chart.title = "Sprint Velocity"
chart.y_axis.title = "Story Points"
chart.x_axis.title = "Sprint"

data = Reference(ws_metrics, min_col=2, max_col=3,
                 min_row=1, max_row=max_sprint+1)
cats = Reference(ws_metrics, min_col=1,
                 min_row=2, max_row=max_sprint+1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_metrics.add_chart(chart, "F2")

# ==========================
# BURNDOWN (for Sprint 1)
# ==========================
ws_burn = wb.create_sheet("Burndown")

ws_burn.append(["Day", "Ideal Remaining", "Actual Remaining"])

total_sp = df[df["Sprint No"] == 1]["Story Points"].sum()

for day in range(1, SPRINT_LENGTH_DAYS + 1):
    ideal = total_sp - (total_sp / SPRINT_LENGTH_DAYS) * day
    actual = total_sp - (day * (total_sp / SPRINT_LENGTH_DAYS) * 0.9)
    ws_burn.append([day, ideal, actual])

line = LineChart()
line.title = "Sprint 1 Burndown"
line.y_axis.title = "Remaining SP"
line.x_axis.title = "Day"

data = Reference(ws_burn, min_col=2, max_col=3,
                 min_row=1, max_row=SPRINT_LENGTH_DAYS+1)
cats = Reference(ws_burn, min_col=1,
                 min_row=2, max_row=SPRINT_LENGTH_DAYS+1)

line.add_data(data, titles_from_data=True)
line.set_categories(cats)
ws_burn.add_chart(line, "E2")

# ==========================
# SAVE OUTPUT
# ==========================
wb.save("Auto_Generated_Sprint_Planner.xlsx")

print("Sprint plan auto-generated from backlog successfully.")